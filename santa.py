import sys, os 
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText

import openpyxl
import random

from jinja2 import Environment, FileSystemLoader, select_autoescape
env = Environment(
  loader=FileSystemLoader(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')),
    autoescape=select_autoescape(['html', 'xml'])
)

# Class: a gifter 
class Santa():

  def __init__(self, name, email, black_list, survey):
    self.name = name 
    self.email = email
    self.black_list = self.set_up_blacklist(black_list, name)
    self.survey = survey
    self.giftee = None 

  def set_up_blacklist(self, blacklist, name):
    if blacklist:
      return blacklist.split(",") + [name]
    else:
      return [name]



# Class: where the work of secret santas takes place 
class NorthPole():

  def generate_santas(self, file):
    print "Generating secret santas!"
    santas = {}
    wb = openpyxl.load_workbook(file)
    ws = wb.active 

    headers = [c.internal_value for c in ws[1]]

    for row in ws.iter_rows(min_row=2):
      cells = [c.internal_value for c in row[0:3]]
      cells.append({headers[i+3]:c.internal_value for i,c in enumerate(row[3:])})
      santa = Santa(*cells)
      santas[santa.name] = santa 
    print "Done."
    return santas 

  def sort_santas(self, santas):
    print "Assigning secret santas..."
    unassigned = set(santas.keys())
    to_match = [(name, set(data.black_list)) for name, data in santas.iteritems()]
    to_match.sort(key=lambda x: len(x[1]))
    to_match.reverse()

    for person, bade in to_match:
      possible = unassigned - bade 
      if len(possible) < 1:
        raise StandardError("Error! Restart")
      else:
        giftee = list(possible)[random.randint(0, len(possible) - 1)]
        unassigned.remove(giftee)
        santas[person].giftee = giftee 

    print "Done."

  def email_santas(self, santas, from_addr, password):
    for name, santa in santas.iteritems():
      print "Emailing %s their secret santa!!" % name 
      giftee_name = santa.giftee 
      giftee = santas[giftee_name]
      self._send_email(santa.email, from_addr, password, giftee.name, giftee.description)
    print "Done."

  # todo: pull html out as separate file 
  def _send_email(self, to_addr, from_addr, password, giftee, email_template, subject='Your Secret Santa!', year=2017):
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(from_addr, password)

    msg = MIMEMultipart()
    msg['From'] = from_addr
    msg['To'] = to_addr
    msg['Subject'] = subject

    body = self._render_template(email_template, giftee)
    
    msg.attach(MIMEText(body, 'html'))
    server.sendmail(from_addr, to_addr, msg.as_string())
    server.quit()

  def _render_template(self, t, giftee, year=2017):
    template = env.get_template(t)
    return template.render(year=year, name=giftee.name, survey=giftee.survey)


if __name__ == '__main__':
  spreadsheet_file = sys.argv[1]  
  email_template = sys.argv[2] 
  email = sys.argv[3]
  password = sys.argv[4]

  np = NorthPole()
  santas = np.generate_santas(spreadsheet_file)
  np.sort_santas(santas)
  np.email_santas(santas, email, password)



